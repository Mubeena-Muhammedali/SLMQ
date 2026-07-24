# -*- coding: utf-8 -*-
import base64
import io
from datetime import datetime, date, timedelta
import pytz

from odoo import models, fields, api, _
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None


DATE_FORMATS = ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%y')


class OdAttendanceTimeoffImportWizard(models.TransientModel):
    _name = 'od.hr.attendance.timeoff.import.wizard'
    _description = 'Import Attendance / Time Off from Timesheet'

    import_file = fields.Binary(string='Timesheet File')
    import_filename = fields.Char(string='File Name')
    error_log = fields.Text(string='Errors', readonly=True)

    # ------------------------------------------------------------------
    # Template download
    # ------------------------------------------------------------------
    def action_download_template(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_url',
            'url': '/hr_attendance_timeoff_import/template',
            'target': 'self',
        }

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _parse_date(self, value):
        """Accepts a date/datetime cell, or common text date formats."""
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if value in (None, ''):
            raise ValueError('empty date')
        text = str(value).strip()
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        raise ValueError('Could not understand date value "%s"' % value)

    @staticmethod
    def _cell(values, index):
        if index is None or index >= len(values):
            return None
        return values[index]

    def _local_to_utc(self, dt, employee):
        """Convert employee local datetime to naive UTC datetime."""

        tz_name = (
            employee.resource_calendar_id.tz
            or employee.user_id.tz
            or self.env.user.tz
            or 'UTC'
        )

        tz = pytz.timezone(tz_name)

        if dt.tzinfo:
            local_dt = dt.astimezone(tz)
        else:
            local_dt = tz.localize(dt)

        utc_dt = local_dt.astimezone(pytz.UTC)

        return utc_dt.replace(tzinfo=None)

    def _get_day_schedule(self, calendar, work_date):
        """Return (standard_hours, start_hour, end_hour, lunch_duration) for
        this employee's calendar on this particular weekday."""
        weekday = str(work_date.weekday())  # Monday=0 ... Sunday=6
        attendance_lines = calendar.attendance_ids.filtered(
            lambda l: l.dayofweek == weekday
        ).sorted(key=lambda l: l.hour_from)

        if attendance_lines:
            work_lines = attendance_lines.filtered(lambda l: l.day_period != 'lunch')
            lunch_lines = attendance_lines.filtered(lambda l: l.day_period == 'lunch')
            standard_hours = sum(l.hour_to - l.hour_from for l in work_lines)
            start_hour = attendance_lines[0].hour_from
            end_hour = attendance_lines[-1].hour_to
            lunch_duration = sum(l.hour_to - l.hour_from for l in lunch_lines)
        else:
            # Not a working day per the calendar (e.g. weekend) - fall back
            # to the company's standard daily hours so a manually entered
            # value can still be processed.
            standard_hours = calendar.hours_per_day or 8.0
            start_hour = 8.0
            end_hour = start_hour + standard_hours
            lunch_duration = 0.0

        return standard_hours, start_hour, end_hour, lunch_duration

    # ------------------------------------------------------------------
    # Main import
    # ------------------------------------------------------------------
    def action_import(self):
        self.ensure_one()

        if not self.import_file:
            raise UserError(_('Please select a file to upload first.'))
        if not openpyxl:
            raise UserError(_('The "openpyxl" Python library is not installed on the server. '
                               'Ask your administrator to run: pip install openpyxl'))

        try:
            content = base64.b64decode(self.import_file)
            workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:
            raise UserError(_('Could not read the uploaded file. Make sure it is a valid '
                               '.xlsx file. Technical error: %s') % exc)

        sheet = workbook.active

        # --- Row 1: column headers ---
        header_cells = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = [str(v).strip().lower() if v else '' for v in header_cells]

        if 'employee name' not in headers:
            raise UserError(_('Row 1 must contain an "Employee Name" column header.'))
        if 'date' not in headers:
            raise UserError(_('Row 1 must contain a "Date" column header.'))

        emp_col = headers.index('employee name')
        date_col = headers.index('date')
        hours_col = headers.index('hours') if 'hours' in headers else None
        timeoff_type_col = headers.index('time off type') if 'time off type' in headers else None
        timeoff_hours_col = headers.index('time off') if 'time off' in headers else None

        if hours_col is None and timeoff_type_col is None:
            raise UserError(_('Row 1 must contain a "Hours" and/or a "Time Off Type" column header.'))

        errors = []
        attendance_vals = []
        leave_vals = []
        employee_cache = {}
        leave_type_cache = {}

        for row in sheet.iter_rows(min_row=2):
            row_num = row[0].row
            values = [c.value for c in row]
            if all(v in (None, '') for v in values):
                continue

            emp_name = self._cell(values, emp_col)
            day_value = self._cell(values, date_col)
            if not emp_name or day_value in (None, ''):
                continue

            # --- resolve employee (cached, so multiple employees/rows are fine) ---
            emp_key = str(emp_name).strip().lower()
            employee = employee_cache.get(emp_key)
            if employee is None:
                found = self.env['hr.employee'].sudo().search(
                    [('name', 'ilike', str(emp_name).strip())])
                if not found:
                    errors.append(_('Row %s: no employee found matching "%s".') % (row_num, emp_name))
                    employee_cache[emp_key] = False
                    continue
                if len(found) > 1:
                    errors.append(_(
                        'Row %s: multiple employees match "%s": %s. Please use a more specific name.'
                    ) % (row_num, emp_name, ', '.join(found.mapped('name'))))
                    employee_cache[emp_key] = False
                    continue
                if not found.resource_calendar_id:
                    errors.append(_('Row %s: employee %s has no working schedule.') % (row_num, found.name))
                    employee_cache[emp_key] = False
                    continue
                employee = found
                employee_cache[emp_key] = employee
            elif employee is False:
                continue

            calendar = employee.resource_calendar_id

            # --- resolve date ---
            try:
                work_date = self._parse_date(day_value)
            except ValueError:
                errors.append(_('Row %s: invalid Date value "%s".') % (row_num, day_value))
                continue

            standard_hours, start_hour, end_hour, lunch_duration = self._get_day_schedule(calendar, work_date)

            hours_value = self._cell(values, hours_col)
            timeoff_type_value = self._cell(values, timeoff_type_col)
            timeoff_hours_value = self._cell(values, timeoff_hours_col)

            has_hours = hours_value not in (None, '')
            has_leave_type = timeoff_type_value not in (None, '')
            has_timeoff_hours = timeoff_hours_value not in (None, '')

            if not has_hours and not has_leave_type:
                continue

            try:
                H = float(hours_value) if has_hours else None
            except Exception:
                errors.append(_('Row %s (day %s): "%s" is not a valid number of Hours.')
                               % (row_num, work_date, hours_value))
                continue

            try:
                T = float(timeoff_hours_value) if has_timeoff_hours else None
            except Exception:
                errors.append(_('Row %s (day %s): "%s" is not a valid number for Time off.')
                               % (row_num, work_date, timeoff_hours_value))
                continue

            leave_type = False
            if has_leave_type:
                type_name = str(timeoff_type_value).strip()
                key = type_name.lower()
                leave_type = leave_type_cache.get(key)
                if leave_type is None:
                    leave_type = self.env['hr.leave.type'].search(
                        [('name', 'ilike', type_name)], limit=1)
                    if not leave_type:
                        errors.append(_(
                            'Row %s (day %s): no Time Off Type found matching "%s".'
                        ) % (row_num, work_date, type_name))
                        continue
                    leave_type_cache[key] = leave_type

            # ------------------------------------------------------------
            # Reconcile Hours worked vs Time off hours against the
            # employee's standard hours for the day:
            #   - Both filled  -> Hours is authoritative; Time off is
            #                     recalculated as (standard - Hours) so the
            #                     day always adds up to the standard hours.
            #   - Only Time Off Type filled -> use the "Time off" value if
            #                     given (e.g. 8 = a full day off), otherwise
            #                     assume a full standard day off. Any hours
            #                     left over (standard - leave) are logged
            #                     as regular attendance automatically.
            #   - Only Hours filled -> plain attendance, no cap (covers
            #                     normal/overtime days).
            # ------------------------------------------------------------
            attendance_hours = 0.0
            leave_hours = 0.0

            if leave_type:
                if has_hours:
                    attendance_hours = H
                    leave_hours = max(0.0, standard_hours - H)
                else:
                    leave_hours = T if has_timeoff_hours else standard_hours
                    leave_hours = max(0.0, min(leave_hours, standard_hours))
                    attendance_hours = max(0.0, standard_hours - leave_hours)
            else:
                attendance_hours = H

            # --- Attendance record ---
            if attendance_hours > 0:
                local_check_in = (
                    datetime.combine(work_date, datetime.min.time())
                    + timedelta(hours=start_hour)
                )
                local_check_out = local_check_in + timedelta(
                    hours=attendance_hours + lunch_duration
                )
                check_in = self._local_to_utc(local_check_in, employee)
                check_out = self._local_to_utc(local_check_out, employee)

                attendance_vals.append({
                    'employee_id': employee.id,
                    'check_in': check_in,
                    'check_out': check_out,
                })

            # --- Leave record ---
            if leave_hours > 0 and leave_type:
                is_full_day = attendance_hours <= 0 and leave_hours >= standard_hours

                if is_full_day:
                    # No attendance that day - leave spans the whole
                    # working day (e.g. Roshan: 8-to-5 fully off).
                    leave_start_hour = start_hour
                    leave_end_hour = end_hour
                else:
                    # Leave starts right where attendance for the day
                    # left off (e.g. worked 8-to-1, so leave is 1-to-5),
                    # instead of overlapping the same start time.
                    leave_start_hour = start_hour + attendance_hours + (
                        lunch_duration if attendance_hours > 0 else 0
                    )
                    leave_end_hour = leave_start_hour + leave_hours

                local_date_from = (
                    datetime.combine(work_date, datetime.min.time())
                    + timedelta(hours=leave_start_hour)
                )
                local_date_to = (
                    datetime.combine(work_date, datetime.min.time())
                    + timedelta(hours=leave_end_hour)
                )

                date_from = self._local_to_utc(local_date_from, employee)
                date_to = self._local_to_utc(local_date_to, employee)

                leave_val = {
                    'employee_id': employee.id,
                    'holiday_status_id': leave_type.id,
                    'request_date_from': work_date,
                    'request_date_to': work_date,
                    'date_from': date_from,
                    'date_to': date_to,
                }
                if not is_full_day:
                    # Partial day off - flag it as an hour-based request so
                    # it displays correctly (e.g. "3 hours" instead of 1 day).
                    if 'request_unit_hours' in self.env['hr.leave']._fields:
                        leave_val.update({
                            'request_unit_hours': True,
                            'request_hour_from': leave_start_hour,
                            'request_hour_to': leave_end_hour,
                        })
                leave_vals.append(leave_val)

        created_attendances = self.env['hr.attendance']
        if attendance_vals:
            try:
                created_attendances = self.env['hr.attendance'].sudo().create(attendance_vals)
            except Exception as exc:
                errors.append(_("Could not create attendance records: %s") % exc)

        created_leaves = self.env['hr.leave']

        for vals in leave_vals:
            try:
                leave = self.env['hr.leave'].sudo().create(vals)
                leave.action_approve()

                created_leaves |= leave

            except Exception as exc:
                emp = self.env['hr.employee'].browse(vals.get('employee_id'))
                errors.append(
                    _("Could not create time off for %s on day %s: %s")
                    % (emp.name, vals.get('request_date_from'), exc)
                )

        self.write({
            'error_log': '\n'.join(errors) if errors else False,
        })
        if errors:

            return {
                'type': 'ir.actions.act_window',
                'res_model': 'od.hr.attendance.timeoff.import.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
            }
        else:
            return True
