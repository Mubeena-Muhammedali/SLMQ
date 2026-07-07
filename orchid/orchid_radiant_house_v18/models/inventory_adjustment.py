# -*- coding: utf-8 -*-

from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64
from io import BytesIO
from openpyxl import load_workbook

class OdInventoryAdjustment(models.Model):
    _name = "od.inventory.adjustment"
    _inherit = ['mail.thread']
    _description = "Inventory Adjustment"

    name = fields.Char(string="Name", default="NEW", tracking=True)
    date = fields.Date(string="Date", default=fields.Date.context_today, tracking=True)
    location_id = fields.Many2one('stock.location', string="Location", tracking=True)
    account_id = fields.Many2one('account.account', string="Account", tracking=True)
    remarks = fields.Char(string="Remarks")
    line_ids = fields.One2many('od.inventory.adjustment.line','od_inventory_id', string="Lines")
    company_id = fields.Many2one('res.company', string="Company", default=lambda self:self.env.company)
    user_id = fields.Many2one('res.users', string="User", default=lambda self:self.env.user)
    state = fields.Selection([('draft','Draft'),('done','Done')], default='draft', string="State", tracking=True)
    data_file = fields.Binary(string='Data', default=False)
    file_name = fields.Char(string='Data File',readonly=True)

    def read_xl_file(self):
        # Decode the uploaded file
        file_content = base64.b64decode(self.data_file)
        workbook = load_workbook(filename=BytesIO(file_content), data_only=True)
        sheet = workbook.active

        values_sheet = []
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if i == 1:
                continue  # Skip header
            row_values = [str(cell).strip() if cell is not None else '' for cell in row]
            values_sheet.append(row_values)
        return values_sheet


    def upload_data(self):
        for record in self:
            if not record.data_file:
                raise UserError(_('Please upload a file first.'))

            value_data = record.read_xl_file()

            if not value_data:
                raise UserError(_('Nothing to upload. File is empty or invalid.'))

            # Optional: remove old lines
            if record.line_ids:
                record.line_ids.unlink()

            line_vals_list = []

            for value in value_data:
                product_code = value[0]
                product = self.env['product.product'].search([('default_code', '=', str(product_code))], limit=1)
                if not product:
                    raise UserError(_("No product found with code '%s'") % product_code)

                # Define the mapping
                line_vals = {
                    'product_id': product.id,
                    'counted_qty': value[1],
                }

                line_vals_list.append((0, 0, line_vals))

            # Create the lines
            if line_vals_list:
                record.line_ids.unlink()
                record.line_ids = line_vals_list
                record.line_ids.od_get_onhand_qty()
            else:
                raise UserError(_("No data!!"))


    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            vals['name'] = self.env['ir.sequence'].next_by_code('od.inventory.adjustment')
        return super(OdInventoryAdjustment, self).create(vals_list)

    def create_adjustment(self):
        if not self.line_ids:
            raise UserError(_("No Lines!!!"))
        for line in self.line_ids:
            quant_vals={
            'location_id':self.location_id.id,
            'product_id':line.product_id.id,
            'inventory_quantity':line.counted_qty,
            'lot_id':line.lot_id and line.lot_id.id or False,
            # 'inventory_diff_quantity':line.differnce_qty,
            }
            quant_id = self.env['stock.quant'].with_context(inventory_mode=True).create(quant_vals)
            type='input'
            if line.differnce_qty<0:
                type='output'
            quant_id.with_context(od_adjustment=True,od_type=type,od_adjustment_account_id=self.account_id).action_apply_inventory()
            line.adjustment_line_id =quant_id.id
        self.state='done'

class OdInventoryAdjustmentLine(models.Model):
    _name = "od.inventory.adjustment.line"
    _description = "Adjustment Lines"

    od_inventory_id = fields.Many2one('od.inventory.adjustment', string="Inventory Adjustment", copy=False, ondelete="cascade")
    product_id = fields.Many2one('product.product', string="Product")
    lot_id = fields.Many2one('stock.lot', string="Lot/Serial No")
    on_hand_qty = fields.Float(string="On Hand")
    counted_qty = fields.Float(string="Counted Qty")
    differnce_qty = fields.Float(string="Difference", compute="_compute_inventory_diff_quantity")
    adjustment_line_id = fields.Many2one('stock.quant', string="Adjustment Line ID")
    tracking = fields.Selection(related='product_id.tracking', readonly=True)
    location_id = fields.Many2one('stock.location', string="Location", related="od_inventory_id.location_id")

    @api.depends('counted_qty','on_hand_qty')
    def _compute_inventory_diff_quantity(self):
        for quant in self:
            quant.differnce_qty = quant.counted_qty - quant.on_hand_qty

    @api.onchange('product_id','location_id','lot_id')
    def od_get_onhand_qty(self):
        for move in self:
            qty = 0
            if move.product_id and move.location_id:
                qty=move.product_id.with_context(location=move.location_id.id,lot_id=move.lot_id.id).qty_available
            move.on_hand_qty = qty
            print("ouiiiiiiiiiiiiiiiiiiii")

class ProductTemplate(models.Model):
    _inherit = 'product.template'

    def _get_product_accounts(self):
        product_accounts = super()._get_product_accounts()

        if self._context.get('od_adjustment') and self._context.get('od_adjustment_account_id') and self._context.get('od_type'):
            if self._context.get('od_type') == 'output':
                product_accounts['stock_output'] = self._context.get('od_adjustment_account_id')
            if self._context.get('od_type') == 'input':
                product_accounts['stock_input'] = self._context.get('od_adjustment_account_id')
        return product_accounts









